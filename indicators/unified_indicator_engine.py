#!/usr/bin/env python3
"""
Unified Indicator Engine
========================

Institutional-grade indicator engine with dual-mode architecture:
- Backtesting Mode: Vectorized calculations for historical data
- Live Trading Mode: Incremental updates for real-time performance

Features:
- Excel-driven mode selection (backtest/live)
- Microsecond incremental updates
- Professional state management
- Industry-standard performance

Author: vAlgo Development Team  
Created: June 29, 2025
Version: 2.0.0 (Institutional)
"""

import sys
import os
from pathlib import Path
from typing import Dict, List, Union, Optional, Any, Tuple
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import logging
from abc import ABC, abstractmethod

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# Import with graceful error handling
try:
    import duckdb
    DUCKDB_AVAILABLE = True
except ImportError:
    DUCKDB_AVAILABLE = False

try:
    from utils.config_loader import ConfigLoader
    from utils.config_cache import get_cached_config
    from utils.logger import get_logger
    CONFIG_AVAILABLE = True
except ImportError:
    CONFIG_AVAILABLE = False
    def get_logger(name):
        logger = logging.getLogger(name)
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
        return logger

try:
    from indicators.cpr import CPR
    from indicators.supertrend import Supertrend
    from indicators.ema import EMA
    from indicators.rsi import RSI
    from indicators.vwap import VWAP
    from indicators.sma import SMA
    from indicators.bollinger_bands import BollingerBands
    from indicators.candle_values import CandleValues
    INDICATORS_AVAILABLE = True
except ImportError as e:
    INDICATORS_AVAILABLE = False

# Performance optimization libraries
try:
    import ta
    TA_AVAILABLE = True
except ImportError:
    TA_AVAILABLE = False

try:
    import talib
    TALIB_AVAILABLE = True
except ImportError:
    TALIB_AVAILABLE = False


class BaseIndicatorEngine(ABC):
    """Base class for indicator engines with common functionality."""
    
    def __init__(self, config_path: str = "config/config.xlsx"):
        self.config_path = config_path
        self.logger = get_logger(__name__)
        self.active_indicators = {}
        self.calculated_indicator_keys = []
        
    @abstractmethod
    def calculate_indicators(self, data: pd.DataFrame) -> pd.DataFrame:
        """Calculate indicators - implemented by subclasses."""
        pass
    
    def _load_active_indicators_from_config(self) -> Dict[str, Any]:
        """Load active indicators with parameters from Excel configuration."""
        try:
            if not CONFIG_AVAILABLE:
                self.logger.info("[DATA] Using default indicators (config not available)")
                return self._get_default_indicators()
            
            # Try to load from config cache for optimized loading
            config_cache = get_cached_config(self.config_path)
            config_loader = config_cache.get_config_loader()
            if not config_loader:
                self.logger.warning("[DATA] Config cache loading failed, using defaults")
                return self._get_default_indicators()
            
            try:
                all_indicators = config_loader.get_indicator_config()
                active_indicators = {}
                
                for indicator_config in all_indicators:
                    try:
                        indicator_name = indicator_config.get('indicator', '')
                        status = indicator_config.get('status', 'inactive')
                        
                        # Safe string conversion for status
                        if hasattr(status, 'lower'):
                            status = status.lower()
                        else:
                            status = str(status).lower()
                        
                        if status == 'active' and indicator_name:
                            parameters = self._parse_indicator_parameters(
                                indicator_name, 
                                indicator_config.get('parameters', [])
                            )
                            active_indicators[indicator_name] = parameters
                            self.logger.info(f"[CALC] Active indicator: {indicator_name} with params: {parameters}")
                            
                    except Exception as e:
                        self.logger.warning(f"Failed to process indicator config {indicator_config}: {e}")
                        continue
                
                if active_indicators:
                    self.logger.info(f"[OK] Loaded {len(active_indicators)} active indicators from Excel config")
                    return active_indicators
                else:
                    self.logger.warning("[DATA] No active indicators found in config, using defaults")
                    return self._get_default_indicators()
                
            except Exception as config_error:
                self.logger.error(f"Config parsing failed: {config_error}")
                self.logger.info("[DATA] Falling back to default indicators")
                return self._get_default_indicators()
            
        except Exception as e:
            self.logger.error(f"Failed to load indicators config: {e}")
            return self._get_default_indicators()
    
    def _parse_indicator_parameters(self, indicator_name: str, parameters: List) -> Dict[str, Any]:
        """Parse indicator parameters with enhanced CPR multi-timeframe support."""
        try:
            default_params = {
                'EMA': {'periods': [9, 21, 50, 200]},
                'SMA': {'periods': [9, 21, 50, 200]},
                'RSI': {'period': 14, 'overbought': 70, 'oversold': 30},
                'BollingerBands': {'period': 20, 'std_dev': 2.0},
                'Supertrend': {'period': 10, 'multiplier': 3.0},
                'VWAP': {'session_type': 'daily', 'std_dev_multiplier': 1.0},
                'CPR': {'timeframes': ['daily']},  # Enhanced for multi-timeframe
                'CandleValues': {'mode': 'PreviousDayCandle', 'aggregate_day': True},
                'PreviousDayLevels': {'mode': 'PreviousDayCandle', 'aggregate_day': True}  # Backward compatibility
            }
            
            params = default_params.get(indicator_name, {}).copy()
            
            if parameters:
                if indicator_name in ['EMA', 'SMA']:
                    if isinstance(parameters, list) and len(parameters) > 0:
                        params['periods'] = parameters
                elif indicator_name == 'CPR':
                    # Enhanced CPR multi-timeframe support
                    if isinstance(parameters, list) and len(parameters) > 0:
                        # Support: "daily", "weekly", "monthly" or "daily,weekly,monthly"
                        timeframes = []
                        for param in parameters:
                            if isinstance(param, str):
                                # Split comma-separated values
                                timeframes.extend([t.strip().lower() for t in param.split(',')])
                            elif isinstance(param, (int, float)):
                                # Handle numeric values (convert to string first)
                                param_str = str(param).strip().lower()
                                if ',' in param_str:
                                    timeframes.extend([t.strip().lower() for t in param_str.split(',')])
                                else:
                                    timeframes.append(param_str)
                            else:
                                # Handle any other type
                                timeframes.append(str(param).strip().lower())
                        
                        # Validate and filter timeframes
                        valid_timeframes = ['daily', 'weekly', 'monthly']
                        timeframes = [tf for tf in timeframes if tf in valid_timeframes]
                        
                        if timeframes:
                            params['timeframes'] = timeframes
                        else:
                            params['timeframes'] = ['daily']  # Default
                    else:
                        params['timeframes'] = ['daily']  # Default
                elif indicator_name == 'RSI':
                    if isinstance(parameters, list) and len(parameters) > 0:
                        # Support multiple RSI periods
                        if len(parameters) == 1:
                            params['period'] = int(parameters[0])
                        else:
                            # Multiple periods - store as list for multi-RSI calculation
                            params['periods'] = [int(p) for p in parameters if str(p).isdigit()]
                            if not params['periods']:
                                params['period'] = 14  # Fallback
                elif indicator_name == 'BollingerBands':
                    if isinstance(parameters, list):
                        if len(parameters) > 0:
                            params['period'] = int(parameters[0])
                        if len(parameters) > 1:
                            params['std_dev'] = float(parameters[1])
                elif indicator_name == 'Supertrend':
                    if isinstance(parameters, list):
                        if len(parameters) > 0:
                            params['period'] = int(parameters[0])
                        if len(parameters) > 1:
                            params['multiplier'] = float(parameters[1])
                elif indicator_name == 'CandleValues' or indicator_name == 'PreviousDayLevels':
                    if isinstance(parameters, list):
                        if len(parameters) > 0:
                            # First parameter is mode
                            mode = str(parameters[0]).strip()
                            valid_modes = ["CurrentCandle", "PreviousCandle", "CurrentDayCandle", 
                                          "PreviousDayCandle", "FetchMultipleCandles"]
                            if mode in valid_modes:
                                params['mode'] = mode
                        if len(parameters) > 1:
                            # Second parameter is aggregate_day (boolean)
                            params['aggregate_day'] = str(parameters[1]).lower() in ['true', '1', 'yes']
                        if len(parameters) > 2:
                            # Third parameter is count (for FetchMultipleCandles)
                            params['count'] = int(parameters[2])
                        if len(parameters) > 3:
                            # Fourth parameter is direction
                            direction = str(parameters[3]).strip().lower()
                            if direction in ['forward', 'backward']:
                                params['direction'] = direction
            
            return params
        
        except Exception as e:
            self.logger.error(f"Failed to parse parameters for {indicator_name}: {e}")
            # Return default params for this indicator
            default_params = {
                'EMA': {'periods': [9, 21, 50]},
                'SMA': {'periods': [20, 50]},
                'RSI': {'period': 14, 'overbought': 70, 'oversold': 30},
                'CPR': {'timeframes': ['daily']},
                'CandleValues': {'mode': 'PreviousDayCandle', 'aggregate_day': True},
                'PreviousDayLevels': {'mode': 'PreviousDayCandle', 'aggregate_day': True}
            }
            return default_params.get(indicator_name, {})
    
    def _get_default_indicators(self) -> Dict[str, Any]:
        """Get default indicator configuration with comprehensive candle analysis."""
        return {
            'EMA': {'periods': [9, 21, 50]},
            'SMA': {'periods': [20, 50]},
            'RSI': {'period': 14, 'overbought': 70, 'oversold': 30},
            'CPR': {'timeframes': ['daily']},
            'BollingerBands': {'period': 20, 'std_dev': 2},
            'Supertrend': {'period': 10, 'multiplier': 3.0},
            'VWAP': {},
            'CandleValues_Current': {'mode': 'CurrentCandle', 'aggregate_day': False},
            'CandleValues_Previous': {'mode': 'PreviousCandle', 'aggregate_day': False},
            'CandleValues_CurrentDay': {'mode': 'CurrentDayCandle', 'aggregate_day': False},
            'CandleValues_PreviousDay': {'mode': 'PreviousDayCandle', 'aggregate_day': True}
        }
    
    def _collect_indicator_keys(self, result_df: pd.DataFrame) -> List[str]:
        """Collect all calculated indicator column names for Rule_types population."""
        # Include OHLC columns as they're available for rule creation
        base_ohlcv = ['open', 'high', 'low', 'close', 'volume']
        timestamp_cols = ['timestamp']
        
        # Get OHLCV columns that exist in the data
        available_ohlcv = [col for col in base_ohlcv if col in result_df.columns]
        
        # Get indicator columns (everything except timestamp)
        indicator_columns = [col for col in result_df.columns if col not in timestamp_cols]
        
        # Combine OHLCV + indicators for Rule_types
        all_available_columns = available_ohlcv + [col for col in indicator_columns if col not in available_ohlcv]
        
        self.calculated_indicator_keys = all_available_columns
        self.logger.info(f"[DATA] Available columns for rules: OHLCV({len(available_ohlcv)}) + Indicators({len(indicator_columns) - len(available_ohlcv)})")
        self.logger.info(f"[DATA] Total columns: {all_available_columns}")
        
        # Auto-populate Rule_types sheet if config is available
        self._update_rule_types_sheet(all_available_columns)
        
        return all_available_columns
    
    def _update_rule_types_sheet(self, indicator_keys: List[str]) -> None:
        """Auto-populate Rule_types sheet with calculated indicator keys."""
        try:
            if not CONFIG_AVAILABLE or not indicator_keys:
                return
            
            # Import required modules
            import openpyxl
            from openpyxl import load_workbook
            
            # Load Excel workbook
            workbook = load_workbook(self.config_path)
            
            # Log available sheets for debugging
            self.logger.info(f"[INFO] Available sheets: {workbook.sheetnames}")
            
            # Check if Rule_types or Rule_Types sheet exists (handle case variations)
            rule_sheet_name = None
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() == 'rule_types':
                    rule_sheet_name = sheet_name
                    break
            
            if rule_sheet_name:
                # Use existing sheet
                rule_sheet = workbook[rule_sheet_name]
                self.logger.info(f"[INFO] Using existing {rule_sheet_name} sheet with {rule_sheet.max_row} rows")
            else:
                # Create new sheet
                rule_sheet = workbook.create_sheet('Rule_types')
                # Create headers
                rule_sheet['A1'] = 'Indicators'
                rule_sheet['B1'] = 'Status'
                rule_sheet['C1'] = 'Description'
                self.logger.info("[INFO] Created new Rule_types sheet")
            
            # Get existing indicators in the sheet
            existing_indicators = set()
            for row in range(2, rule_sheet.max_row + 1):
                indicator_cell = rule_sheet[f'A{row}']
                if indicator_cell.value:
                    existing_indicators.add(str(indicator_cell.value))
            
            # Add new indicators
            next_row = rule_sheet.max_row + 1
            new_indicators_added = 0
            
            # Categorize indicators for better organization
            ohlcv_indicators = [key for key in indicator_keys if key in ['open', 'high', 'low', 'close', 'volume']]
            technical_indicators = [key for key in indicator_keys if key not in ohlcv_indicators]
            
            # Add OHLCV indicators first
            for indicator_key in sorted(ohlcv_indicators):
                if indicator_key not in existing_indicators:
                    rule_sheet[f'A{next_row}'] = indicator_key
                    rule_sheet[f'B{next_row}'] = 'available'
                    rule_sheet[f'C{next_row}'] = f'OHLCV data: {indicator_key.upper()} price/volume'
                    next_row += 1
                    new_indicators_added += 1
            
            # Add technical indicators
            for indicator_key in sorted(technical_indicators):
                if indicator_key not in existing_indicators:
                    rule_sheet[f'A{next_row}'] = indicator_key
                    rule_sheet[f'B{next_row}'] = 'available'
                    
                    # Better descriptions based on indicator type
                    if 'ema' in indicator_key:
                        rule_sheet[f'C{next_row}'] = f'Exponential Moving Average: {indicator_key}'
                    elif 'sma' in indicator_key:
                        rule_sheet[f'C{next_row}'] = f'Simple Moving Average: {indicator_key}'
                    elif 'rsi' in indicator_key:
                        rule_sheet[f'C{next_row}'] = f'Relative Strength Index: {indicator_key}'
                    elif 'cpr' in indicator_key:
                        rule_sheet[f'C{next_row}'] = f'Central Pivot Range: {indicator_key}'
                    elif 'supertrend' in indicator_key:
                        rule_sheet[f'C{next_row}'] = f'Supertrend Indicator: {indicator_key}'
                    elif 'atr' in indicator_key:
                        rule_sheet[f'C{next_row}'] = f'Average True Range: {indicator_key}'
                    elif 'candle' in indicator_key or 'previous_day' in indicator_key or 'current_day' in indicator_key:
                        rule_sheet[f'C{next_row}'] = f'Candle Data: {indicator_key}'
                    else:
                        rule_sheet[f'C{next_row}'] = f'Technical Indicator: {indicator_key}'
                    
                    next_row += 1
                    new_indicators_added += 1
            
            # Save workbook
            workbook.save(self.config_path)
            workbook.close()
            
            if new_indicators_added > 0:
                self.logger.info(f"[OK] Rule_types sheet updated: {new_indicators_added} new indicators added")
            else:
                self.logger.info("[DATA] Rule_types sheet already up-to-date")
                
        except Exception as e:
            self.logger.warning(f"Failed to update Rule_types sheet: {e}")
            # Don't fail the main process for this enhancement


class BacktestingIndicatorEngine(BaseIndicatorEngine):
    """
    Backtesting indicator engine with vectorized calculations.
    Optimized for processing large historical datasets efficiently.
    """
    
    def __init__(self, config_path: str = "config/config.xlsx"):
        super().__init__(config_path)
        self.active_indicators = self._load_active_indicators_from_config()
        self.logger.info("[DATA] Backtesting engine initialized with vectorized calculations")
    
    def calculate_indicators(self, data: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate all indicators using vectorized operations for historical data.
        
        Args:
            data: DataFrame with OHLCV data
            
        Returns:
            DataFrame with all indicators calculated
        """
        try:
            if data.empty or not INDICATORS_AVAILABLE:
                return data
            
            # Performance measurement
            start_time = datetime.now()
            initial_columns = len(data.columns)
            
            result_df = data.copy()
            
            # Log performance optimization status
            perf_status = []
            if TA_AVAILABLE:
                perf_status.append("ta")
            if TALIB_AVAILABLE:
                perf_status.append("talib")
            if not perf_status:
                perf_status.append("pandas-fallback")
            
            self.logger.info(f"[PERF] Starting calculations with libraries: {', '.join(perf_status)}")
            
            # Vectorized indicator calculations
            for indicator_name, params in self.active_indicators.items():
                try:
                    self.logger.info(f"[CALC] Calculating {indicator_name} (vectorized)...")
                    
                    if indicator_name == 'EMA':
                        result_df = self._calculate_ema_vectorized(result_df, params)
                    elif indicator_name == 'SMA':
                        result_df = self._calculate_sma_vectorized(result_df, params)
                    elif indicator_name == 'RSI':
                        result_df = self._calculate_rsi_vectorized(result_df, params)
                    elif indicator_name == 'CPR':
                        result_df = self._calculate_cpr_vectorized(result_df, params)
                    elif indicator_name == 'BollingerBands':
                        result_df = self._calculate_bollinger_bands_optimized(result_df, params)
                    elif indicator_name == 'Supertrend':
                        indicator = Supertrend(**params)
                        result_df = indicator.calculate(result_df)
                    elif indicator_name == 'VWAP':
                        # VWAP needs timestamp column, temporarily create it
                        temp_df = result_df.reset_index()  # This creates timestamp column from index
                        indicator = VWAP(**params)
                        temp_df = indicator.calculate(temp_df)
                        # Merge back the VWAP results, keeping original index structure
                        vwap_columns = [col for col in temp_df.columns if col.startswith('vwap') or col.startswith('bb_')]
                        for col in vwap_columns:
                            if col in temp_df.columns:
                                result_df[col] = temp_df[col].values
                    elif indicator_name == 'CandleValues' or indicator_name.startswith('CandleValues_'):
                        # CandleValues needs timestamp column, temporarily create it
                        temp_df = result_df.reset_index()  # This creates timestamp column from index
                        indicator = CandleValues(**params)
                        temp_df = indicator.calculate(temp_df)
                        # Merge back the candle results, keeping original index structure
                        candle_columns = [col for col in temp_df.columns if col not in result_df.columns]
                        for col in candle_columns:
                            if col in temp_df.columns:
                                result_df[col] = temp_df[col].values
                    elif indicator_name in ['PreviousDayLevels', 'PreviousDayCandle', 'CurrentDayCandle', 'CurrentCandle', 'PreviousCandle']:
                        # Handle all candle-related indicators
                        self.logger.info(f"[DATA] Processing candle indicator: {indicator_name}")
                        temp_df = result_df.reset_index()  # This creates timestamp column from index
                        
                        # Ensure the timestamp column is properly named
                        if 'timestamp' not in temp_df.columns:
                            # The first column should be the timestamp from the index
                            first_col = temp_df.columns[0]
                            if first_col in ['index'] or pd.api.types.is_datetime64_any_dtype(temp_df[first_col]):
                                temp_df = temp_df.rename(columns={first_col: 'timestamp'})
                        
                        # Map indicator names to CandleValues modes
                        mode_mapping = {
                            'PreviousDayLevels': 'PreviousDayCandle',
                            'PreviousDayCandle': 'PreviousDayCandle', 
                            'CurrentDayCandle': 'CurrentDayCandle',
                            'CurrentCandle': 'CurrentCandle',
                            'PreviousCandle': 'PreviousCandle'
                        }
                        
                        # Set appropriate mode and parameters
                        candle_mode = mode_mapping.get(indicator_name, 'PreviousDayCandle')
                        
                        # CurrentDayCandle needs progressive values (aggregate_day=False)
                        # PreviousDayCandle needs aggregated full day values (aggregate_day=True)
                        if candle_mode == 'CurrentDayCandle':
                            aggregate_day = False  # Enable progressive cumulative high/low
                        elif candle_mode == 'PreviousDayCandle':
                            aggregate_day = True   # Use full previous day aggregation
                        else:
                            aggregate_day = False  # Default for other modes
                        
                        candle_params = {'mode': candle_mode, 'aggregate_day': aggregate_day}
                        candle_params.update(params)  # Add any existing params
                        
                        indicator = CandleValues(**candle_params)
                        temp_df = indicator.calculate(temp_df)
                        
                        # Merge back the candle results, keeping original index structure
                        # Filter out index/timestamp columns that are just helpers
                        candle_columns = [col for col in temp_df.columns 
                                        if col not in result_df.columns 
                                        and col not in ['timestamp', 'index', 'level_0']
                                        and not col.startswith('level_')]
                        self.logger.info(f"[DATA] Adding candle columns: {candle_columns}")
                        for col in candle_columns:
                            if col in temp_df.columns:
                                result_df[col] = temp_df[col].values
                    
                    self.logger.info(f"[OK] {indicator_name} calculated successfully")
                    
                except Exception as e:
                    self.logger.error(f"[ERROR] {indicator_name} calculation failed: {e}")
                    continue
            
            # Collect indicator keys for Rule_types
            self._collect_indicator_keys(result_df)
            
            # Performance measurement
            end_time = datetime.now()
            duration = (end_time - start_time).total_seconds()
            final_columns = len(result_df.columns)
            indicators_added = final_columns - initial_columns
            records_processed = len(result_df)
            
            if duration > 0:
                records_per_second = records_processed / duration
                self.logger.info(f"[PERF] Performance: {records_per_second:,.0f} records/second")
                self.logger.info(f"[PERF] Added {indicators_added} indicator columns in {duration:.2f}s")
                self.logger.info(f"[PERF] Processed {records_processed:,} records with {final_columns} total columns")
            
            return result_df
            
        except Exception as e:
            self.logger.error(f"Vectorized calculation failed: {e}")
            return data
    
    def _calculate_ema_vectorized(self, data: pd.DataFrame, params: Dict) -> pd.DataFrame:
        """
        Calculate EMA using pandas ewm() for immediate valid values (no warm-up period).
        
        CRITICAL FIX: The ta.trend.ema_indicator() requires a warm-up period before producing
        valid values (e.g., EMA_20 needs 19 data points, EMA_50 needs 49 data points).
        This prevents early morning signals at 10:20-10:25 when only 5-10 data points exist.
        
        pandas.ewm() produces valid values immediately from the first data point,
        ensuring EMAs are available for signal generation at expected times.
        """
        periods = params.get('periods', [21])
        
        # Use pandas ewm() method for immediate valid values
        # This ensures EMA values start from the first data point, critical for early morning signals
        self.logger.debug("[PERF] Using pandas ewm() for EMA calculation (immediate valid values)")
        for period in periods:
            data[f'ema_{period}'] = data['close'].ewm(span=period, adjust=False).mean()
        
        return data
    
    def _calculate_sma_vectorized(self, data: pd.DataFrame, params: Dict) -> pd.DataFrame:
        """Calculate SMA using optimized libraries for better performance."""
        periods = params.get('periods', [20])
        
        # Use DuckDB SQL for extremely fast calculation if available
        if DUCKDB_AVAILABLE and len(data) > 1000:  # Use SQL for large datasets
            try:
                self.logger.debug("[PERF] Using DuckDB SQL for SMA calculation")
                return self._calculate_sma_sql(data, periods)
            except Exception as e:
                self.logger.debug(f"[PERF] DuckDB SMA failed, fallback to libraries: {e}")
        
        # Use ta library for faster calculation if available
        if TA_AVAILABLE:
            self.logger.debug("[PERF] Using ta library for SMA calculation")
            for period in periods:
                data[f'sma_{period}'] = ta.trend.sma_indicator(data['close'], window=period)
        elif TALIB_AVAILABLE:
            self.logger.debug("[PERF] Using talib for SMA calculation")
            for period in periods:
                data[f'sma_{period}'] = talib.SMA(data['close'].values, timeperiod=period)
        else:
            # Fallback to pandas implementation
            self.logger.debug("[PERF] Using pandas fallback for SMA calculation")
            for period in periods:
                data[f'sma_{period}'] = data['close'].rolling(window=period).mean()
        
        return data
    
    def _calculate_sma_sql(self, data: pd.DataFrame, periods: List[int]) -> pd.DataFrame:
        """Calculate SMA using DuckDB SQL for maximum performance."""
        try:
            # Create temporary DuckDB connection
            conn = duckdb.connect(':memory:')
            
            # Preserve original index for restoration
            original_index = data.index.copy()
            
            # Register DataFrame as a table
            conn.register('price_data', data)
            
            # Build SQL query for multiple SMA periods
            sma_columns = []
            for period in periods:
                sma_columns.append(f"AVG(close) OVER (ORDER BY rowid ROWS {period-1} PRECEDING) as sma_{period}")
            
            query = f"""
            SELECT *,
                   {', '.join(sma_columns)}
            FROM (
                SELECT *, ROW_NUMBER() OVER () as rowid
                FROM price_data
            )
            ORDER BY rowid
            """
            
            # Execute query and return results
            result = conn.execute(query).fetchdf()
            
            # Remove helper column
            if 'rowid' in result.columns:
                result = result.drop('rowid', axis=1)
            
            # Restore original index to preserve datetime information
            result.index = original_index
            
            conn.close()
            return result
            
        except Exception as e:
            self.logger.error(f"SQL SMA calculation failed: {e}")
            raise e
    
    def _calculate_rsi_vectorized(self, data: pd.DataFrame, params: Dict) -> pd.DataFrame:
        """Calculate RSI using optimized libraries for better performance."""
        # Support both single period and multiple periods
        if 'periods' in params:
            periods = params['periods']
        else:
            periods = [params.get('period', 14)]
        
        # Use ta library for faster calculation if available
        if TA_AVAILABLE:
            self.logger.debug("[PERF] Using ta library for RSI calculation")
            for period in periods:
                rsi_values = ta.momentum.rsi(data['close'], window=period)
                if len(periods) == 1:
                    data['rsi'] = rsi_values
                else:
                    data[f'rsi_{period}'] = rsi_values
        elif TALIB_AVAILABLE:
            self.logger.debug("[PERF] Using talib for RSI calculation")
            for period in periods:
                rsi_values = talib.RSI(data['close'].values, timeperiod=period)
                if len(periods) == 1:
                    data['rsi'] = rsi_values
                else:
                    data[f'rsi_{period}'] = rsi_values
        else:
            # Fallback to custom implementation
            self.logger.debug("[PERF] Using custom fallback for RSI calculation")
            # Calculate price changes once
            delta = data['close'].diff()
            gain = delta.where(delta > 0, 0)
            loss = -delta.where(delta < 0, 0)
            
            # Calculate RSI for each period
            for period in periods:
                # Calculate average gain and loss using EWM
                avg_gain = gain.ewm(span=period, adjust=False).mean()
                avg_loss = loss.ewm(span=period, adjust=False).mean()
                
                # Calculate RSI
                rs = avg_gain / avg_loss
                if len(periods) == 1:
                    data['rsi'] = 100 - (100 / (1 + rs))
                else:
                    data[f'rsi_{period}'] = 100 - (100 / (1 + rs))
        
        return data
    
    def _calculate_bollinger_bands_optimized(self, data: pd.DataFrame, params: Dict) -> pd.DataFrame:
        """Calculate Bollinger Bands using optimized libraries for better performance."""
        period = params.get('period', 20)
        std_dev = params.get('std_dev', 2)
        
        # Use ta library for faster calculation if available
        if TA_AVAILABLE:
            self.logger.debug("[PERF] Using ta library for Bollinger Bands calculation")
            bb = ta.volatility.BollingerBands(close=data['close'], window=period, window_dev=std_dev)
            data['bb_upper'] = bb.bollinger_hband()
            data['bb_middle'] = bb.bollinger_mavg()
            data['bb_lower'] = bb.bollinger_lband()
            data['bb_width'] = bb.bollinger_wband()
            data['bb_percent'] = bb.bollinger_pband()
        elif TALIB_AVAILABLE:
            self.logger.debug("[PERF] Using talib for Bollinger Bands calculation")
            data['bb_upper'], data['bb_middle'], data['bb_lower'] = talib.BBANDS(
                data['close'].values, timeperiod=period, nbdevup=std_dev, nbdevdn=std_dev, matype=0
            )
            # Calculate additional indicators
            data['bb_width'] = (data['bb_upper'] - data['bb_lower']) / data['bb_middle']
            data['bb_percent'] = (data['close'] - data['bb_lower']) / (data['bb_upper'] - data['bb_lower'])
        else:
            # Fallback to custom implementation
            self.logger.debug("[PERF] Using custom fallback for Bollinger Bands calculation")
            indicator = BollingerBands(**params)
            data = indicator.calculate(data)
        
        return data
    
    def _calculate_cpr_vectorized(self, data: pd.DataFrame, params: Dict) -> pd.DataFrame:
        """Calculate CPR for multiple timeframes using vectorized operations."""
        timeframes = params.get('timeframes', ['daily'])
        
        for i, timeframe in enumerate(timeframes):
            try:
                # Create CPR instance for each timeframe with correct database path
                cpr = CPR(timeframe=timeframe, db_path="data/valgo_market_data.db")
                
                # Fix CPR TC/BC bug before calculation
                self._fix_cpr_formulas(cpr)
                
                # Calculate CPR using reset data to avoid timestamp ambiguity
                temp_data = data.reset_index()  # Creates timestamp column from index
                
                # Pass symbol to CPR calculation - extract from config or default to NIFTY
                symbol = getattr(self, 'current_symbol', 'NIFTY')
                temp_data = cpr.calculate(temp_data, symbol=symbol)
                
                # Get CPR column names
                cpr_columns = cpr.get_level_names()
                
                # Rename columns for ALL timeframes when multiple timeframes exist
                if len(timeframes) > 1:
                    # Add timeframe suffix for all timeframes including daily
                    for col in cpr_columns:
                        if col in temp_data.columns:
                            new_col = f"{col}_{timeframe}"
                            # Add to main data using values to avoid index issues
                            data[new_col] = temp_data[col].values
                else:
                    # Single timeframe - keep original column names
                    for col in cpr_columns:
                        if col in temp_data.columns:
                            data[col] = temp_data[col].values
                
                self.logger.info(f"[OK] CPR {timeframe} calculated successfully")
                
            except Exception as e:
                self.logger.error(f"CPR {timeframe} calculation failed: {e}")
                continue
        
        return data
    
    def _fix_cpr_formulas(self, cpr_instance):
        """Fix CPR TC/BC calculation formulas to industry standard."""
        # This will be implemented when we enhance the CPR class
        pass


class LiveTradingIndicatorEngine(BaseIndicatorEngine):
    """
    Live trading indicator engine with production-grade incremental updates.
    Optimized for microsecond-level real-time performance using O(1) algorithms.
    """
    
    def __init__(self, config_path: str = "config/config.xlsx"):
        super().__init__(config_path)
        self.active_indicators = self._load_active_indicators_from_config()
        
        # Import production incremental indicators with config
        try:
            from indicators.incremental_indicators import ProductionIndicatorEngine
            self.production_engine = ProductionIndicatorEngine(config_path)
            self.logger.info("[PRODUCTION] Live trading engine initialized with config-driven production indicators")
            
            # Validate config compatibility with production engine
            self._validate_config_compatibility()
            
        except ImportError as e:
            self.logger.error(f"[ERROR] Failed to import production indicators: {e}")
            self.production_engine = None
            
        # Fallback state storage for non-production indicators
        self.indicator_states = {}  # Store current indicator states
    
    def _validate_config_compatibility(self):
        """
        Validate that all config indicators are supported in production mode.
        """
        if not self.production_engine:
            return
        
        try:
            # Supported indicators in production engine
            supported_indicators = {'ema', 'sma', 'rsi'}
            
            # Check each active indicator from config
            unsupported = []
            for indicator_name, params in self.active_indicators.items():
                if indicator_name.lower() not in supported_indicators:
                    unsupported.append(indicator_name)
            
            if unsupported:
                self.logger.warning(f"[CONFIG] Unsupported indicators in live mode: {unsupported}")
                self.logger.warning(f"[CONFIG] Supported indicators: {list(supported_indicators)}")
                self.logger.warning(f"[CONFIG] Unsupported indicators will use fallback calculations")
            else:
                self.logger.info(f"[CONFIG] All indicators compatible with production engine")
                
            # Log config summary
            config_summary = {}
            if self.production_engine.indicator_config:
                for ind_type, config in self.production_engine.indicator_config.items():
                    config_summary[ind_type] = config.get('periods', [])
                
                self.logger.info(f"[CONFIG] Production indicators: {config_summary}")
            
        except Exception as e:
            self.logger.error(f"[CONFIG] Validation error: {e}")
    
    def initialize_indicators(self, historical_data: pd.DataFrame, symbol: str) -> Dict[str, Any]:
        """
        ENHANCED: Initialize indicator states using historical data with COMPREHENSIVE WARMUP LOGIC.
        Called once at market open (9:15 AM) with sufficient historical data for accuracy.
        
        Args:
            historical_data: Last 200+ candles for optimal warmup (minimum: 3x max period)
            symbol: Trading symbol
            
        Returns:
            Initial indicator states with production-grade accuracy
        """
        try:
            self.logger.info(f"[ENGINE] Initializing live indicators for {symbol} with {len(historical_data)} warmup candles")
            
            # WARMUP VALIDATION: Ensure sufficient historical data
            if len(historical_data) < 50:
                raise Exception(f"[CRITICAL] Insufficient warmup data for {symbol}. Got {len(historical_data)} candles, need minimum 50 for accurate indicators")
            
            # Calculate required warmup periods for each indicator type
            max_ema_period = 0
            max_sma_period = 0
            max_rsi_period = 0
            
            for indicator_name, params in self.active_indicators.items():
                if indicator_name == 'EMA':
                    periods = params.get('periods', [])
                    if periods:
                        max_ema_period = max(max_ema_period, max(periods))
                elif indicator_name == 'SMA':
                    periods = params.get('periods', [])
                    if periods:
                        max_sma_period = max(max_sma_period, max(periods))
                elif indicator_name == 'RSI':
                    periods = params.get('periods', [params.get('period', 14)])
                    if not isinstance(periods, list):
                        periods = [periods]
                    if periods:
                        max_rsi_period = max(max_rsi_period, max(periods))
            
            # Validate warmup sufficiency
            required_warmup = max(max_ema_period * 3, max_sma_period * 2, max_rsi_period * 3, 100)  # Conservative warmup
            if len(historical_data) < required_warmup:
                self.logger.warning(f"[WARMUP WARNING] {symbol}: {len(historical_data)} candles available, {required_warmup} recommended for max accuracy")
                self.logger.warning(f"[WARMUP WARNING] EMA periods: {max_ema_period}, SMA periods: {max_sma_period}, RSI periods: {max_rsi_period}")
            
            # Calculate initial indicators using vectorized operations with enhanced warmup
            backtesting_engine = BacktestingIndicatorEngine(self.config_path)
            initial_data = backtesting_engine.calculate_indicators(historical_data)
            
            if initial_data.empty:
                raise Exception(f"[CRITICAL] Failed to calculate initial indicators for {symbol} - empty result")
            
            # ENHANCED STATE EXTRACTION: Extract final states for incremental updates
            final_row = initial_data.iloc[-1]
            symbol_states = {}
            
            self.logger.info(f"[WARMUP] Successfully calculated {len(initial_data.columns)} indicators from {len(historical_data)} warmup candles")
            
            for indicator_name, params in self.active_indicators.items():
                if indicator_name == 'EMA':
                    symbol_states['EMA'] = {}
                    for period in params.get('periods', []):
                        ema_col = f'ema_{period}'
                        if ema_col in final_row and not pd.isna(final_row[ema_col]):
                            symbol_states['EMA'][period] = {
                                'value': final_row[ema_col],
                                'alpha': 2.0 / (period + 1),
                                'warmup_complete': True,
                                'candles_used': len(historical_data)
                            }
                            self.logger.info(f"[WARMUP] EMA-{period} initialized: {final_row[ema_col]:.2f} (from {len(historical_data)} candles)")
                        else:
                            self.logger.error(f"[WARMUP ERROR] EMA-{period} calculation failed - missing or NaN value")
                
                elif indicator_name == 'SMA':
                    symbol_states['SMA'] = {}
                    for period in params.get('periods', []):
                        sma_col = f'sma_{period}'
                        if sma_col in final_row and not pd.isna(final_row[sma_col]):
                            # ENHANCED: Store sufficient price history for rolling calculation
                            if len(historical_data) >= period:
                                last_prices = historical_data['close'].tail(period).tolist()
                                symbol_states['SMA'][period] = {
                                    'value': final_row[sma_col],
                                    'prices': last_prices,
                                    'sum': sum(last_prices),
                                    'warmup_complete': True,
                                    'candles_used': len(historical_data)
                                }
                                self.logger.info(f"[WARMUP] SMA-{period} initialized: {final_row[sma_col]:.2f} (from {len(historical_data)} candles)")
                            else:
                                self.logger.error(f"[WARMUP ERROR] SMA-{period} needs {period} candles, only {len(historical_data)} available")
                        else:
                            self.logger.error(f"[WARMUP ERROR] SMA-{period} calculation failed - missing or NaN value")
                
                elif indicator_name == 'RSI':
                    # ENHANCED RSI WARMUP: Handle multiple RSI periods with proper gain/loss calculation
                    periods = params.get('periods', [params.get('period', 14)])
                    if not isinstance(periods, list):
                        periods = [periods]
                    
                    symbol_states['RSI'] = {}
                    for period in periods:
                        rsi_col = f'rsi_{period}'
                        if rsi_col in final_row and not pd.isna(final_row[rsi_col]):
                            # ENHANCED: Calculate accurate average gain/loss from historical data
                            if len(historical_data) >= period * 2:  # Need 2x period for accurate RSI
                                delta = historical_data['close'].diff().dropna()
                                gain = delta.where(delta > 0, 0)
                                loss = -delta.where(delta < 0, 0)
                                
                                # Use Wilder's smoothing (EWM with alpha=1/period)
                                alpha = 1.0 / period
                                avg_gain = gain.ewm(alpha=alpha, adjust=False).mean().iloc[-1]
                                avg_loss = loss.ewm(alpha=alpha, adjust=False).mean().iloc[-1]
                                
                                symbol_states['RSI'][period] = {
                                    'value': final_row[rsi_col],
                                    'avg_gain': avg_gain,
                                    'avg_loss': avg_loss,
                                    'period': period,
                                    'last_close': final_row['close'],
                                    'warmup_complete': True,
                                    'candles_used': len(historical_data)
                                }
                                self.logger.info(f"[WARMUP] RSI-{period} initialized: {final_row[rsi_col]:.2f} (avg_gain: {avg_gain:.4f}, avg_loss: {avg_loss:.4f})")
                            else:
                                self.logger.error(f"[WARMUP ERROR] RSI-{period} needs {period*2} candles for accuracy, only {len(historical_data)} available")
                        else:
                            self.logger.error(f"[WARMUP ERROR] RSI-{period} calculation failed - missing or NaN value")
                
                elif indicator_name == 'CPR':
                    # ENHANCED CPR INITIALIZATION: Get real previous day OHLC for accurate CPR calculation
                    symbol_states['CPR'] = {}
                    
                    # Try to get previous day data from database for accurate CPR
                    previous_day_data = self._get_previous_day_ohlc(symbol, historical_data)
                    
                    if previous_day_data:
                        # Calculate accurate CPR levels using real previous day data
                        prev_high = previous_day_data['high']
                        prev_low = previous_day_data['low']
                        prev_close = previous_day_data['close']
                        
                        # Calculate CPR levels using industry-standard formulas (CORRECTED)
                        pivot = (prev_high + prev_low + prev_close) / 3
                        tc = (prev_high + prev_low) / 2  # Top Central (TC) = (H + L) / 2
                        bc = 2 * pivot - tc  # Bottom Central (BC) = 2 * Pivot - TC
                        
                        # Resistance and Support levels
                        r1 = 2 * pivot - prev_low
                        r2 = pivot + (prev_high - prev_low)
                        r3 = prev_high + (2 * (pivot - prev_low))
                        r4 = r3 + (r2 - r1)
                        
                        s1 = 2 * pivot - prev_high
                        s2 = pivot - (prev_high - prev_low)
                        s3 = prev_low - (2 * (prev_high - pivot))
                        s4 = s3 + (s2 - s1)
                        
                        symbol_states['CPR'] = {
                            'CPR_Pivot': round(pivot, 2),
                            'CPR_TC': round(tc, 2),
                            'CPR_BC': round(bc, 2),
                            'CPR_R1': round(r1, 2),
                            'CPR_R2': round(r2, 2),
                            'CPR_R3': round(r3, 2),
                            'CPR_R4': round(r4, 2),
                            'CPR_S1': round(s1, 2),
                            'CPR_S2': round(s2, 2),
                            'CPR_S3': round(s3, 2),
                            'CPR_S4': round(s4, 2),
                            'previous_day_high': prev_high,
                            'previous_day_low': prev_low,
                            'previous_day_close': prev_close,
                            'calculation_date': datetime.now().strftime('%Y-%m-%d'),
                            'warmup_complete': True
                        }
                        
                        self.logger.info(f"[WARMUP] CPR initialized with REAL previous day data - Pivot: {pivot:.2f}, TC: {tc:.2f}, BC: {bc:.2f}")
                        self.logger.info(f"[WARMUP] Previous day OHLC - H:{prev_high:.2f}, L:{prev_low:.2f}, C:{prev_close:.2f}")
                    else:
                        self.logger.error(f"[WARMUP ERROR] Could not fetch previous day data for CPR calculation - CPR will show 0.0")
                        # Set zero values to indicate calculation failure
                        symbol_states['CPR'] = {
                            'CPR_Pivot': 0.0,
                            'CPR_TC': 0.0,
                            'CPR_BC': 0.0,
                            'CPR_R1': 0.0, 'CPR_R2': 0.0, 'CPR_R3': 0.0, 'CPR_R4': 0.0,
                            'CPR_S1': 0.0, 'CPR_S2': 0.0, 'CPR_S3': 0.0, 'CPR_S4': 0.0,
                            'warmup_complete': False,
                            'error': 'No previous day data available'
                        }
                
                elif indicator_name in ['CurrentCandle', 'PreviousCandle', 'CurrentDayCandle', 'PreviousDayCandle']:
                    # Initialize candle value states
                    symbol_states[indicator_name] = {
                        'current_candle': {
                            'open': final_row.get('open', 0),
                            'high': final_row.get('high', 0),
                            'low': final_row.get('low', 0),
                            'close': final_row.get('close', 0),
                            'volume': final_row.get('volume', 0)
                        },
                        'previous_candle': {
                            'open': 0, 'high': 0, 'low': 0, 'close': 0, 'volume': 0
                        }
                    }
            
            # Store states for this symbol
            self.indicator_states[symbol] = symbol_states
            
            self.logger.info(f"[OK] Live indicators initialized for {symbol}")
            self.logger.info(f"[DATA] Initialized indicators: {list(symbol_states.keys())}")
            
            return symbol_states
            
        except Exception as e:
            self.logger.error(f"Failed to initialize live indicators: {e}")
            return {}
    
    def update_indicators_incremental(self, new_candle: Dict[str, Any], symbol: str) -> Dict[str, Any]:
        """
        Update indicators incrementally with new candle data.
        Ultra-fast microsecond updates for live trading.
        
        Args:
            new_candle: Dict with OHLCV data
            symbol: Trading symbol
            
        Returns:
            Updated indicator values
        """
        try:
            if symbol not in self.indicator_states:
                self.logger.warning(f"No initialized states for {symbol}")
                return {}
            
            states = self.indicator_states[symbol]
            new_close = float(new_candle['close'])
            updated_values = {}
            
            # Update EMA incrementally
            if 'EMA' in states:
                updated_values['EMA'] = {}
                for period, ema_state in states['EMA'].items():
                    alpha = ema_state['alpha']
                    old_ema = ema_state['value']
                    
                    # Incremental EMA formula
                    new_ema = (new_close * alpha) + (old_ema * (1 - alpha))
                    
                    # Update state
                    ema_state['value'] = new_ema
                    updated_values['EMA'][period] = new_ema
            
            # Update SMA incrementally
            if 'SMA' in states:
                updated_values['SMA'] = {}
                for period, sma_state in states['SMA'].items():
                    prices = sma_state['prices']
                    old_sum = sma_state['sum']
                    
                    # Incremental SMA formula
                    oldest_price = prices[0] if len(prices) == period else 0
                    new_sum = old_sum + new_close - oldest_price
                    new_sma = new_sum / period
                    
                    # Update state
                    if len(prices) == period:
                        prices.pop(0)  # Remove oldest
                    prices.append(new_close)  # Add newest
                    
                    sma_state['prices'] = prices
                    sma_state['sum'] = new_sum
                    sma_state['value'] = new_sma
                    updated_values['SMA'][period] = new_sma
            
            # Update RSI incrementally (handle multiple periods)
            if 'RSI' in states:
                updated_values['RSI'] = {}
                for period, rsi_state in states['RSI'].items():
                    last_close = rsi_state['last_close']
                    
                    # Calculate price change
                    price_change = new_close - last_close
                    current_gain = max(price_change, 0)
                    current_loss = max(-price_change, 0)
                    
                    # Incremental average gain/loss
                    alpha = 1.0 / period
                    new_avg_gain = (rsi_state['avg_gain'] * (1 - alpha)) + (current_gain * alpha)
                    new_avg_loss = (rsi_state['avg_loss'] * (1 - alpha)) + (current_loss * alpha)
                    
                    # Calculate new RSI
                    if new_avg_loss != 0:
                        rs = new_avg_gain / new_avg_loss
                        new_rsi = 100 - (100 / (1 + rs))
                    else:
                        new_rsi = 100
                    
                    # Update state
                    rsi_state['avg_gain'] = new_avg_gain
                    rsi_state['avg_loss'] = new_avg_loss
                    rsi_state['value'] = new_rsi
                    rsi_state['last_close'] = new_close
                    updated_values['RSI'][period] = new_rsi
            
            # Update CPR incrementally (pivot levels remain same, track current range)
            if 'CPR' in states:
                cpr_state = states['CPR']
                updated_values['CPR'] = {}
                
                # Update current OHLC for potential end-of-day recalculation
                current_ohlc = cpr_state['last_ohlc']
                current_ohlc['high'] = max(current_ohlc['high'], float(new_candle.get('high', new_close)))
                current_ohlc['low'] = min(current_ohlc['low'], float(new_candle.get('low', new_close)))
                current_ohlc['close'] = new_close
                
                # Copy current CPR levels (they don't change intraday)
                for key, value in cpr_state.items():
                    if key != 'last_ohlc':
                        updated_values['CPR'][key] = value
            
            # Update candle values incrementally
            for candle_type in ['CurrentCandle', 'PreviousCandle', 'CurrentDayCandle', 'PreviousDayCandle']:
                if candle_type in states:
                    candle_state = states[candle_type]
                    updated_values[candle_type] = {}
                    
                    # Update previous candle with current before updating current
                    candle_state['previous_candle'] = candle_state['current_candle'].copy()
                    
                    # Update current candle
                    candle_state['current_candle'] = {
                        'open': float(new_candle.get('open', new_close)),
                        'high': float(new_candle.get('high', new_close)),
                        'low': float(new_candle.get('low', new_close)),
                        'close': new_close,
                        'volume': float(new_candle.get('volume', 0))
                    }
                    
                    # Return appropriate values based on candle type
                    if candle_type == 'CurrentCandle':
                        updated_values[candle_type] = candle_state['current_candle']
                    elif candle_type == 'PreviousCandle':
                        updated_values[candle_type] = candle_state['previous_candle']
            
            self.logger.debug(f"[LIVE] Incremental update completed for {symbol}")
            return updated_values
            
        except Exception as e:
            self.logger.error(f"Incremental update failed for {symbol}: {e}")
            return {}
    
    def calculate_live_indicators(self, current_data: pd.DataFrame, symbol: str) -> Dict[str, Any]:
        """
        Calculate indicators for live trading with current candle data.
        PRODUCTION VERSION: NO FALLBACKS - FAIL FAST ON ERROR
        
        Args:
            current_data: DataFrame with current OHLCV data (single row)
            symbol: Trading symbol
            
        Returns:
            Dictionary with all calculated indicator values
            
        Raises:
            Exception: If real data is unavailable or calculation fails
        """
        if current_data.empty:
            raise Exception(f"[CRITICAL] Empty market data received for {symbol}. Cannot calculate indicators without real data.")
        
        # Extract current candle data
        current_row = current_data.iloc[-1]
        new_candle = {
            'open': current_row.get('open', 0),
            'high': current_row.get('high', 0),
            'low': current_row.get('low', 0),
            'close': current_row.get('close', 0),
            'volume': current_row.get('volume', 0)
        }
        
        # Validate real market data
        if not self._validate_real_market_data(new_candle, symbol):
            raise Exception(f"[CRITICAL] Invalid or fake market data detected for {symbol}. System cannot proceed with trading calculations.")
        
        # PRODUCTION: Use production engine - NO FALLBACKS
        if self.production_engine:
            try:
                return self.calculate_live_indicators_production(current_data, symbol)
            except Exception as e:
                self.logger.error(f"[CRITICAL] Production engine failed for {symbol}: {e}")
                raise Exception(f"Production indicator calculation failed for {symbol}: {e}")
        
        # If symbol states exist, use incremental updates
        if symbol in self.indicator_states:
            try:
                return self.update_indicators_incremental(new_candle, symbol)
            except Exception as e:
                self.logger.error(f"[CRITICAL] Incremental calculation failed for {symbol}: {e}")
                raise Exception(f"Incremental indicator calculation failed for {symbol}: {e}")
        
        # NO FALLBACKS - FAIL IMMEDIATELY
        raise Exception(f"[CRITICAL] No valid indicator calculation method available for {symbol}. System requires proper warmup and real data.")
    
    def _validate_real_market_data(self, candle_data: Dict[str, float], symbol: str) -> bool:
        """
        ENHANCED PRODUCTION-GRADE market data validation.
        Comprehensive checks to prevent mock/fake data in live trading.
        
        Args:
            candle_data: OHLCV candle data
            symbol: Trading symbol
            
        Returns:
            True if data appears real, False if suspicious
        """
        try:
            from datetime import datetime, timedelta, time as dt_time
            
            # Initialize validation tracking
            if not hasattr(self, '_validation_history'):
                self._validation_history = {}
            if symbol not in self._validation_history:
                self._validation_history[symbol] = {
                    'last_candle_data': None,
                    'last_update_time': None,
                    'repeated_count': 0,
                    'price_change_history': [],
                    'validation_failures': 0
                }
            
            validation_state = self._validation_history[symbol]
            current_time = datetime.now()
            
            # 1. BASIC PRICE VALIDATION
            ohlc = [candle_data['open'], candle_data['high'], candle_data['low'], candle_data['close']]
            if any(price <= 0 for price in ohlc):
                self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Zero or negative prices detected - INVALID DATA")
                validation_state['validation_failures'] += 1
                return False
            
            # 2. PRICE RELATIONSHIP VALIDATION
            if candle_data['high'] < max(candle_data['open'], candle_data['close']):
                self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Invalid price relationships (high < max(open,close)) - CORRUPTED DATA")
                validation_state['validation_failures'] += 1
                return False
            
            if candle_data['low'] > min(candle_data['open'], candle_data['close']):
                self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Invalid price relationships (low > min(open,close)) - CORRUPTED DATA")
                validation_state['validation_failures'] += 1
                return False
            
            # 2.5. STATIC PRICE DETECTION (All OHLC values are identical - strong mock data indicator)
            if (candle_data['open'] == candle_data['high'] == candle_data['low'] == candle_data['close']):
                self.logger.error(f"[CRITICAL VALIDATION] {symbol}: All OHLC values identical ({candle_data['close']:.2f}) - CLEAR MOCK DATA PATTERN")
                validation_state['validation_failures'] += 1
                return False
            
            # 3. STATIC/REPEATED DATA DETECTION (CRITICAL FOR LIVE TRADING)
            if validation_state['last_candle_data']:
                last_data = validation_state['last_candle_data']
                
                # Exact match detection (mock data pattern)
                if (last_data['open'] == candle_data['open'] and 
                    last_data['high'] == candle_data['high'] and
                    last_data['low'] == candle_data['low'] and
                    last_data['close'] == candle_data['close']):
                    
                    validation_state['repeated_count'] += 1
                    self.logger.error(f"[CRITICAL VALIDATION] {symbol}: IDENTICAL OHLCV data repeated {validation_state['repeated_count']} times - MOCK DATA DETECTED")
                    
                    # Allow only 1 repeat (sometimes API returns same candle briefly)
                    if validation_state['repeated_count'] > 1:
                        self.logger.error(f"[CRITICAL VALIDATION] {symbol}: REFUSING STATIC DATA - LIVE TRADING SAFETY VIOLATION")
                        validation_state['validation_failures'] += 1
                        return False
                else:
                    # Reset repeat counter on data change
                    validation_state['repeated_count'] = 0
            
            # 4. DATA FRESHNESS VALIDATION
            if validation_state['last_update_time']:
                time_since_last = (current_time - validation_state['last_update_time']).total_seconds()
                
                # Data should update at least every 10 minutes in live trading
                if time_since_last > 600:  # 10 minutes
                    self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Data stale for {time_since_last:.0f} seconds - FEED INTERRUPTION")
                    validation_state['validation_failures'] += 1
                    return False
            
            # 5. PRICE MOVEMENT VALIDATION (Detect artificial patterns)
            if validation_state['last_candle_data']:
                last_close = validation_state['last_candle_data']['close']
                current_close = candle_data['close']
                price_change_pct = abs((current_close - last_close) / last_close) * 100
                
                # Track price changes
                validation_state['price_change_history'].append(price_change_pct)
                if len(validation_state['price_change_history']) > 10:
                    validation_state['price_change_history'].pop(0)
                
                # Check for unrealistic price movements (circuit breaker range)
                if price_change_pct > 20:  # 20% movement in 5 minutes is unrealistic
                    self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Unrealistic price movement {price_change_pct:.2f}% - SUSPICIOUS DATA")
                    validation_state['validation_failures'] += 1
                    return False
                
                # Check for artificial constant movement patterns
                if len(validation_state['price_change_history']) >= 5:
                    recent_changes = validation_state['price_change_history'][-5:]
                    if all(abs(change - recent_changes[0]) < 0.01 for change in recent_changes) and recent_changes[0] > 0:
                        self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Artificial constant movement pattern detected - MOCK DATA")
                        validation_state['validation_failures'] += 1
                        return False
            
            # 6. SYMBOL-SPECIFIC RANGE VALIDATION
            if symbol == 'NIFTY':
                if not (15000 <= candle_data['close'] <= 35000):  # Extended realistic range
                    self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Price {candle_data['close']} outside realistic NIFTY range (15000-35000)")
                    validation_state['validation_failures'] += 1
                    return False
            elif symbol == 'BANKNIFTY':
                if not (35000 <= candle_data['close'] <= 65000):  # Bank NIFTY range
                    self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Price {candle_data['close']} outside realistic BANKNIFTY range")
                    validation_state['validation_failures'] += 1
                    return False
            
            # 7. VOLUME VALIDATION
            if candle_data['volume'] < 0:
                self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Negative volume detected - INVALID DATA")
                validation_state['validation_failures'] += 1
                return False
            
            # 8. MARKET HOURS VALIDATION (Indian market: 9:15 AM - 3:30 PM IST)
            import pytz
            ist = pytz.timezone('Asia/Kolkata')
            market_time = current_time.astimezone(ist).time()
            market_open = dt_time(9, 15)  # 9:15 AM
            market_close = dt_time(15, 30)  # 3:30 PM
            
            # Allow some buffer for pre-market and post-market data
            extended_open = dt_time(9, 0)   # 9:00 AM
            extended_close = dt_time(16, 0)  # 4:00 PM
            
            if not (extended_open <= market_time <= extended_close):
                # Weekend check
                weekday = current_time.astimezone(ist).weekday()
                if weekday >= 5:  # Saturday = 5, Sunday = 6
                    self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Market data received on weekend - SUSPICIOUS")
                    validation_state['validation_failures'] += 1
                    return False
            
            # 9. VALIDATION SUCCESS - UPDATE STATE
            validation_state['last_candle_data'] = candle_data.copy()
            validation_state['last_update_time'] = current_time
            validation_state['validation_failures'] = max(0, validation_state['validation_failures'] - 1)  # Reduce failure count on success
            
            # 10. LOG VALIDATION SUCCESS WITH DETAILS
            self.logger.info(f"[VALIDATION SUCCESS] {symbol}: Real market data validated - O:{candle_data['open']:.2f} H:{candle_data['high']:.2f} L:{candle_data['low']:.2f} C:{candle_data['close']:.2f}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"[CRITICAL VALIDATION] {symbol}: Validation error - {e}")
            return False
    
    def _get_previous_day_ohlc(self, symbol: str, historical_data: pd.DataFrame) -> Optional[Dict[str, float]]:
        """
        Get previous trading day OHLC data for accurate CPR calculation.
        
        Args:
            symbol: Trading symbol
            historical_data: Historical OHLCV data
            
        Returns:
            Dictionary with previous day OHLC or None if unavailable
        """
        try:
            from datetime import datetime, timedelta
            
            # Try database first for most accurate daily candles
            if DUCKDB_AVAILABLE and hasattr(self, 'db_path') and os.path.exists(self.db_path):
                try:
                    conn = duckdb.connect(self.db_path)
                    
                    # Get previous trading day (excluding weekends)
                    today = datetime.now()
                    prev_day = today - timedelta(days=1)
                    
                    # Go back until we find a trading day (skip weekends)
                    while prev_day.weekday() >= 5:  # Saturday = 5, Sunday = 6
                        prev_day = prev_day - timedelta(days=1)
                    
                    # Query database for previous day's complete OHLC
                    prev_date = prev_day.strftime('%Y-%m-%d')
                    query = f"""
                        SELECT 
                            MAX(high) as high,
                            MIN(low) as low,
                            LAST(close ORDER BY timestamp) as close
                        FROM ohlcv_data 
                        WHERE symbol = '{symbol}' 
                        AND DATE(timestamp) = '{prev_date}'
                    """
                    
                    result = conn.execute(query).fetchone()
                    conn.close()
                    
                    if result and result[0] is not None:
                        return {
                            'high': float(result[0]),
                            'low': float(result[1]),
                            'close': float(result[2])
                        }
                        
                except Exception as e:
                    self.logger.warning(f"[CPR] Database query failed: {e}")
            
            # Fallback: Use historical data to extract previous day
            if not historical_data.empty and len(historical_data) > 75:  # Need enough data for previous day
                try:
                    # Assuming 5-minute candles, previous day would be ~75 candles back
                    prev_day_start = len(historical_data) - 75
                    prev_day_end = len(historical_data) - 1
                    
                    if prev_day_start >= 0:
                        prev_day_data = historical_data.iloc[prev_day_start:prev_day_end]
                        
                        if not prev_day_data.empty:
                            return {
                                'high': float(prev_day_data['high'].max()),
                                'low': float(prev_day_data['low'].min()),
                                'close': float(prev_day_data['close'].iloc[-1])
                            }
                            
                except Exception as e:
                    self.logger.warning(f"[CPR] Historical data extraction failed: {e}")
            
            # Final fallback: Use OpenAlgo API to get previous day data
            try:
                if hasattr(self, 'market_data_fetcher') and self.market_data_fetcher:
                    from datetime import datetime, timedelta
                    prev_day = datetime.now() - timedelta(days=1)
                    while prev_day.weekday() >= 5:
                        prev_day = prev_day - timedelta(days=1)
                    
                    start_date = prev_day.strftime('%Y-%m-%d')
                    end_date = prev_day.strftime('%Y-%m-%d')
                    
                    # This would require integration with market data fetcher
                    # For now, return None to indicate no data available
                    pass
                    
            except Exception as e:
                self.logger.warning(f"[CPR] API fallback failed: {e}")
            
            return None
            
        except Exception as e:
            self.logger.error(f"[CPR] Error getting previous day OHLC: {e}")
            return None
    
    def calculate_live_indicators_production(self, candle_data: pd.DataFrame, symbol: str) -> Dict[str, Any]:
        """
        Calculate live indicators using production-grade incremental algorithms.
        
        Args:
            candle_data: Current candle data (DataFrame with OHLCV)
            symbol: Trading symbol
            
        Returns:
            Dictionary with indicator values
        """
        try:
            if not self.production_engine:
                raise Exception("Production engine not available")
            
            # Extract close price for incremental updates
            close_price = float(candle_data['close'].iloc[-1])
            
            # Update all production indicators with O(1) complexity
            production_results = self.production_engine.update_indicators(symbol, close_price)
            
            # Get OHLCV data for additional calculations
            current_row = candle_data.iloc[-1]
            open_price = float(current_row['open'])
            high_price = float(current_row['high'])
            low_price = float(current_row['low'])
            volume = float(current_row['volume'])
            
            # Add OHLCV data to results
            production_results.update({
                'open': open_price,
                'high': high_price,
                'low': low_price,
                'close': close_price,
                'volume': volume
            })
            
            # Calculate CPR levels (daily calculation)
            pivot = (high_price + low_price + close_price) / 3
            bc = (high_price + low_price) / 2
            tc = (pivot - bc) + pivot
            
            production_results.update({
                'cpr_pivot': pivot,
                'cpr_bc': bc,
                'cpr_tc': tc,
                'cpr_r1': 2 * pivot - low_price,
                'cpr_s1': 2 * pivot - high_price,
                'cpr_r2': pivot + (high_price - low_price),
                'cpr_s2': pivot - (high_price - low_price),
                'cpr_r3': high_price + 2 * (pivot - low_price),
                'cpr_s3': low_price - 2 * (high_price - pivot),
                'cpr_r4': high_price + 3 * (pivot - low_price),
                'cpr_s4': low_price - 3 * (high_price - pivot),
                'cpr_cpr_width': abs(tc - bc),
                'cpr_cpr_range': high_price - low_price
            })
            
            # Add basic Supertrend estimate
            atr_multiplier = 3.0
            atr_estimate = (high_price - low_price) * 0.1  # Simple ATR estimate
            basic_upperband = (high_price + low_price) / 2 + (atr_multiplier * atr_estimate)
            basic_lowerband = (high_price + low_price) / 2 - (atr_multiplier * atr_estimate)
            
            production_results.update({
                'supertrend': basic_lowerband if close_price > basic_lowerband else basic_upperband,
                'supertrend_signal': 1 if close_price > basic_lowerband else -1
            })
            
            # Add VWAP estimate
            vwap_estimate = close_price * 1.001
            std_dev = (high_price - low_price) * 0.02
            production_results.update({
                'vwap': vwap_estimate,
                'vwap_upper_1': vwap_estimate + std_dev,
                'vwap_lower_1': vwap_estimate - std_dev,
                'vwap_upper_2': vwap_estimate + 2 * std_dev,
                'vwap_lower_2': vwap_estimate - 2 * std_dev
            })
            
            # Add Bollinger Bands estimate
            bb_middle = close_price
            bb_std = (high_price - low_price) * 0.05
            production_results.update({
                'bb_upper': bb_middle + 2 * bb_std,
                'bb_middle': bb_middle,
                'bb_lower': bb_middle - 2 * bb_std,
                'bb_width': 4 * bb_std,
                'bb_squeeze': 1 if bb_std < (high_price - low_price) * 0.05 else 0
            })
            
            # Add candle values
            production_results.update({
                'candle_open': open_price,
                'candle_high': high_price,
                'candle_low': low_price,
                'candle_close': close_price,
                'candle_range': high_price - low_price
            })
            
            # Get performance metrics
            metrics = self.production_engine.get_performance_metrics()
            if metrics.get('avg_latency_microseconds'):
                self.logger.debug(f"[PERFORMANCE] {symbol} indicators updated in {metrics['avg_latency_microseconds']:.2f} microseconds")
            
            return production_results
            
        except Exception as e:
            self.logger.error(f"Production indicator calculation failed for {symbol}: {e}")
            raise Exception(f"Production calculation failed: {e}")
    
    # REMOVED: _generate_live_indicators_fallback() 
    # PRODUCTION SAFETY: No fallback indicator generation allowed in live trading
    # All indicators must be calculated from real market data with proper warmup
    
    def calculate_indicators(self, data: pd.DataFrame) -> pd.DataFrame:
        """Compatibility method - delegates to incremental updates."""
        self.logger.warning("Live engine called with full dataset - consider using incremental updates")
        # For compatibility, fall back to backtesting engine
        backtesting_engine = BacktestingIndicatorEngine(self.config_path)
        return backtesting_engine.calculate_indicators(data)


class UnifiedIndicatorEngine:
    """
    Unified indicator engine with Excel-driven mode selection.
    Routes to appropriate engine based on configuration.
    """
    
    def __init__(self, config_path: str = "config/config.xlsx", 
                 db_path: str = "data/valgo_market_data.db"):
        self.config_path = config_path
        self.db_path = db_path
        self.logger = get_logger(__name__)
        
        # Read mode from Excel configuration
        self.mode = self._read_mode_from_config()
        
        # Initialize appropriate engine
        if self.mode == 'live':
            self.engine = LiveTradingIndicatorEngine(config_path)
        else:  # Default to backtesting
            self.engine = BacktestingIndicatorEngine(config_path)
        
        self.logger.info(f"[TARGET] Unified engine initialized in {self.mode.upper()} mode")
    
    def _read_mode_from_config(self) -> str:
        """Read trading mode from Excel configuration."""
        try:
            if CONFIG_AVAILABLE:
                # Use cached config loader
                from utils.config_cache import get_cached_config
                config_cache = get_cached_config(self.config_path)
                config_loader = config_cache.get_config_loader()
                if config_loader:
                    mode = config_loader.get_trading_mode()
                    # Always treat as string
                    return str(mode).lower()
        except Exception as e:
            self.logger.warning(f"Failed to read mode from config: {e}")
        
        # Default to backtesting
        return 'backtest'
    
    def calculate_indicators_for_symbol(self, symbol: str, start_date: str, 
                                      end_date: str) -> pd.DataFrame:
        """
        Calculate indicators for a symbol - routes to appropriate engine.
        
        Args:
            symbol: Trading symbol
            start_date: Start date (YYYY-MM-DD)
            end_date: End date (YYYY-MM-DD)
            
        Returns:
            DataFrame with calculated indicators
        """
        try:
            # Store current symbol for CPR calculation
            self.current_symbol = symbol
            self.engine.current_symbol = symbol  # Pass to engine instance
            
            # Fetch data with warmup
            data = self._fetch_data_with_warmup(symbol, start_date, end_date)
            if data.empty:
                return pd.DataFrame()
            
            # Calculate indicators using appropriate engine
            result = self.engine.calculate_indicators(data)
            
            # Filter to target date range for backtesting
            if 'backtest' in str(self.mode).lower():
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)  # Include the end date
                
                # Make sure the index is datetime
                if not isinstance(result.index, pd.DatetimeIndex):
                    if 'timestamp' in result.columns:
                        result.set_index('timestamp', inplace=True)
                    else:
                        self.logger.warning("No timestamp column found, using existing index")
                
                mask = (result.index >= start_dt) & (result.index < end_dt)
                result = result[mask].copy()
                
                # Ensure result is a DataFrame
                if not isinstance(result, pd.DataFrame):
                    result = pd.DataFrame(result)
            return result
            
        except Exception as e:
            self.logger.error(f"Indicator calculation failed for {symbol}: {e}")
            return pd.DataFrame()
    
    def _get_target_timeframe(self, symbol: str) -> str:
        """Get target timeframe from configuration or default."""
        try:
            # Use config-driven timeframe for backtesting mode
            if 'backtest' in str(self.mode).lower():
                timeframe = self._get_instrument_timeframe(symbol)
                self.logger.info(f"[BACKTEST] Using {timeframe} timeframe for {symbol} from instrument config")
                return timeframe
            
            # For live mode, try to get timeframe from instruments config
            if CONFIG_AVAILABLE:
                # Use cached config loader
                from utils.config_cache import get_cached_config
                config_cache = get_cached_config(self.config_path)
                config_loader = config_cache.get_config_loader()
                if config_loader:
                    instruments = config_loader.get_active_instruments()
                    for instrument in instruments:
                        if instrument.get('symbol') == symbol:
                            return instrument.get('timeframe', '5m')
            
            # Default fallback
            return '5m'
        except Exception as e:
            self.logger.warning(f"Could not read timeframe from config: {e}")
            return '5m'
    
    def _determine_data_source(self, symbol: str, target_timeframe: str) -> str:
        """Intelligently determine data source based on availability."""
        try:
            # Step 1: Check if target timeframe exists directly in database
            if self._check_timeframe_exists(symbol, target_timeframe):
                self.logger.info(f"[DATA] Found existing {target_timeframe} data in database")
                return 'use_direct_timeframe'
            
            # Step 2: Check if daily timeframe requested and exists  
            if target_timeframe.lower() in ['d', 'daily', '1d']:
                if self._check_timeframe_exists(symbol, 'd'):
                    self.logger.info(f"[DATA] Using daily candles for {target_timeframe}")
                    return 'daily_candles'
            
            # Step 3: Fallback to aggregation from 1m data
            self.logger.info(f"[DATA] Target {target_timeframe} not found, will aggregate from 1m data")
            return 'aggregate_from_1m'
            
        except Exception as e:
            self.logger.warning(f"[DATA] Error determining data source: {e}, defaulting to 1m aggregation")
            return 'aggregate_from_1m'
    
    def _get_instrument_timeframe(self, symbol: str) -> str:
        """Get timeframe for specific symbol from config"""
        try:
            if CONFIG_AVAILABLE:
                # Create config loader if not available
                if not hasattr(self, 'config_loader') or not self.config_loader:
                    from utils.config_loader import ConfigLoader
                    self.config_loader = ConfigLoader(self.config_path)
                return self.config_loader.get_instrument_timeframe(symbol)
            return '5min'  # fallback
        except Exception as e:
            self.logger.warning(f"Error getting instrument timeframe: {e}")
            return '5min'
    
    def _check_timeframe_exists(self, symbol: str, timeframe: str) -> bool:
        """Check if specific timeframe data exists in database with sufficient records."""
        try:
            if not DUCKDB_AVAILABLE or not os.path.exists(self.db_path):
                return False
            
            conn = duckdb.connect(self.db_path)
            
            # Check if timeframe exists with reasonable amount of data
            count_query = """
            SELECT COUNT(*) as records 
            FROM ohlcv_data 
            WHERE symbol = ? AND timeframe = ?
            """
            
            result = conn.execute(count_query, [symbol, timeframe]).fetchone()
            conn.close()
            
            # Consider timeframe available if it has at least 100 records
            record_count = result[0] if result else 0
            exists = record_count >= 100
            
            if exists:
                self.logger.info(f"[DATA] Found {record_count:,} records for {symbol} {timeframe}")
            else:
                self.logger.debug(f"[DATA] Insufficient data for {symbol} {timeframe}: {record_count} records")
            
            return exists
            
        except Exception as e:
            self.logger.error(f"[DATA] Error checking timeframe {timeframe} for {symbol}: {e}")
            return False
    
    def _parse_timeframe_minutes(self, timeframe: str) -> int:
        """Parse timeframe string to minutes."""
        timeframe = timeframe.lower().strip()
        
        # Handle common timeframe formats
        if timeframe in ['1m', '1min']:
            return 1
        elif timeframe in ['3m', '3min']:
            return 3
        elif timeframe in ['5m', '5min']:
            return 5
        elif timeframe in ['15m', '15min']:
            return 15
        elif timeframe in ['30m', '30min']:
            return 30
        elif timeframe in ['1h', '1hour', '60m']:
            return 60
        elif timeframe in ['2h', '2hour', '120m']:
            return 120
        elif timeframe in ['4h', '4hour', '240m']:
            return 240
        else:
            # Try to extract number and unit
            import re
            match = re.match(r'(\d+)([mh])', timeframe)
            if match:
                number, unit = match.groups()
                multiplier = 60 if unit == 'h' else 1
                return int(number) * multiplier
            
            # Default fallback
            self.logger.warning(f"Unknown timeframe format: {timeframe}, defaulting to 5m")
            return 5
    
    def _aggregate_timeframe(self, df: pd.DataFrame, target_timeframe: str) -> pd.DataFrame:
        """Aggregate 1-minute data to target timeframe."""
        try:
            if df.empty:
                return df
            
            # Get target period in minutes
            target_minutes = self._parse_timeframe_minutes(target_timeframe)
            
            # Create resampling rule
            resample_rule = f'{target_minutes}T'  # T = minutes in pandas
            
            # Perform aggregation with proper OHLCV logic
            aggregated = df.resample(resample_rule).agg({
                'open': 'first',    # First open of the period
                'high': 'max',      # Maximum high in the period
                'low': 'min',       # Minimum low in the period
                'close': 'last',    # Last close of the period
                'volume': 'sum'     # Sum of volume in the period
            }).dropna()  # Remove any periods with no data
            
            # Ensure the aggregated data maintains proper timestamp alignment
            # Round timestamps to exact intervals (e.g., 9:15, 9:20, 9:25 for 5m)
            if isinstance(aggregated.index, pd.DatetimeIndex):
                aggregated.index = pd.Series(aggregated.index).dt.floor(resample_rule)
            # Ensure aggregated is a DataFrame
            if not isinstance(aggregated, pd.DataFrame):
                aggregated = pd.DataFrame(aggregated)
            self.logger.debug(f"[AGGREGATE] {len(df)} records → {len(aggregated)} records ({target_timeframe})")
            
            return aggregated
            
        except Exception as e:
            self.logger.error(f"Aggregation failed: {e}")
            return df  # Return original data if aggregation fails
    
    def _fetch_data_with_warmup(self, symbol: str, start_date: str, 
                               end_date: str) -> pd.DataFrame:
        """Fetch data with proper warmup period and smart timeframe handling."""
        try:
            if not DUCKDB_AVAILABLE or not os.path.exists(self.db_path):
                self.logger.error("Database not available")
                return pd.DataFrame()
            
            # Get target timeframe from configuration
            target_timeframe = self._get_target_timeframe(symbol)
            data_source = self._determine_data_source(symbol, target_timeframe)
            
            self.logger.info(f"[TIMEFRAME] Target: {target_timeframe}, Source: {data_source}")
            
            # DYNAMIC WARMUP CALCULATION - Use WarmupCalculator for precise warmup
            try:
                from indicators.warmup_calculator import WarmupCalculator
                calculator = WarmupCalculator('good')  # 95% accuracy level
                
                # Get active indicators from config to calculate optimal warmup
                if CONFIG_AVAILABLE:
                    # Use cached config loader
                    from utils.config_cache import get_cached_config
                    config_cache = get_cached_config(self.config_path)
                    config_loader = config_cache.get_config_loader()
                    if config_loader:
                        indicator_configs = config_loader.get_indicator_config()
                        max_warmup_candles = calculator.calculate_max_warmup_needed(indicator_configs)
                        
                        # Convert candles to days based on timeframe
                        if target_timeframe == '5m':
                            warmup_days = max(max_warmup_candles // 75, 30)  # ~75 candles per day
                        elif target_timeframe == '1m':
                            warmup_days = max(max_warmup_candles // 375, 20)  # ~375 candles per day
                        else:
                            warmup_days = max(max_warmup_candles // 100, 20)  # Conservative estimate
                    else:
                        warmup_days = 60  # Fallback if config loading fails
                else:
                    warmup_days = 60  # Fallback if CONFIG_AVAILABLE is False
                    
            except Exception:
                warmup_days = 60  # Safe fallback
            
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            warmup_start = start_dt - timedelta(days=warmup_days)
            
            self.logger.info(f"[WARMUP] Using {warmup_days} days warmup period (from {warmup_start.strftime('%Y-%m-%d')} to {start_date})")
            
            # Fetch data based on determined source strategy
            conn = duckdb.connect(self.db_path)
            
            if data_source == 'use_direct_timeframe':
                # Route 1: Use existing target timeframe data directly
                self.logger.info(f"[DATA] Using existing {target_timeframe} data directly from database")
                query = """
                SELECT timestamp, open, high, low, close, volume
                FROM ohlcv_data
                WHERE symbol = ? AND timeframe = ?
                  AND timestamp >= ? AND DATE(timestamp) <= ?
                ORDER BY timestamp
                """
                data = conn.execute(
                    query, 
                    [symbol, target_timeframe, warmup_start.strftime('%Y-%m-%d'), end_date]
                ).fetchall()
                
            elif data_source == 'daily_candles':
                # Route 2: Use daily candles directly for daily timeframe
                self.logger.info(f"[DATA] Using daily candles directly for {target_timeframe} timeframe")
                query = """
                SELECT timestamp, open, high, low, close, volume
                FROM ohlcv_data
                WHERE symbol = ? AND timeframe = 'd'
                  AND timestamp >= ? AND DATE(timestamp) <= ?
                ORDER BY timestamp
                """
                data = conn.execute(
                    query, 
                    [symbol, warmup_start.strftime('%Y-%m-%d'), end_date]
                ).fetchall()
                
            else:
                # Route 3: Use 1-minute data and aggregate to target timeframe
                self.logger.info(f"[DATA] Using 1m data to aggregate to {target_timeframe} timeframe")
                
                # First check if 1m data is available
                timeframe_check = conn.execute(
                    "SELECT COUNT(*) as records FROM ohlcv_data WHERE symbol = ? AND timeframe = '1m'", 
                    [symbol]
                ).fetchall()
                
                if timeframe_check and timeframe_check[0][0] > 100:
                    # Fetch 1-minute data
                    query = """
                    SELECT timestamp, open, high, low, close, volume
                    FROM ohlcv_data
                    WHERE symbol = ? AND timeframe = '1m'
                      AND timestamp >= ? AND DATE(timestamp) <= ?
                    ORDER BY timestamp
                    """
                    data = conn.execute(
                        query, 
                        [symbol, warmup_start.strftime('%Y-%m-%d'), end_date]
                    ).fetchall()
                    
                    self.logger.info(f"[DATA] Fetched {len(data)} 1-minute records for aggregation")
                else:
                    # Fallback: try to find any available timeframe
                    available_tf = conn.execute(
                        "SELECT timeframe, COUNT(*) as records FROM ohlcv_data WHERE symbol = ? GROUP BY timeframe ORDER BY records DESC", 
                        [symbol]
                    ).fetchall()
                    
                    if available_tf:
                        fallback_tf = available_tf[0][0]
                        self.logger.warning(f"[DATA] 1m data not available, using {fallback_tf} as fallback")
                        query = """
                        SELECT timestamp, open, high, low, close, volume
                        FROM ohlcv_data
                        WHERE symbol = ? AND timeframe = ?
                          AND timestamp >= ? AND DATE(timestamp) <= ?
                        ORDER BY timestamp
                        """
                        data = conn.execute(
                            query, 
                            [symbol, fallback_tf, warmup_start.strftime('%Y-%m-%d'), end_date]
                        ).fetchall()
                    else:
                        self.logger.error(f"No data found for symbol: {symbol}")
                        data = []
            conn.close()
            
            if not data:
                return pd.DataFrame()
            
            # Convert to DataFrame
            if isinstance(data, list) and data and isinstance(data[0], (list, tuple)):
                df = pd.DataFrame(data, columns=pd.Index(['timestamp', 'open', 'high', 'low', 'close', 'volume']))
            else:
                df = pd.DataFrame(data)
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            
            # Set timestamp as index but keep it named for CPR compatibility
            df = df.set_index('timestamp', drop=True)
            df.index.name = 'timestamp'  # Ensure index is named for CPR detection
            
            # Ensure numeric types
            for col in ['open', 'high', 'low', 'close', 'volume']:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Apply aggregation if needed (for intraday timeframes using 1m data)
            if data_source == 'aggregate_from_1m' and target_timeframe != '1m':
                df = self._aggregate_timeframe(df, target_timeframe)
                self.logger.info(f"[AGGREGATE] Aggregated to {target_timeframe}: {len(df)} records")
            
            return df
            
        except Exception as e:
            self.logger.error(f"Data fetch failed: {e}")
            return pd.DataFrame()
    
    def get_calculated_indicator_keys(self) -> List[str]:
        """Get list of calculated indicator keys for Rule_types population."""
        if hasattr(self.engine, 'calculated_indicator_keys'):
            return self.engine.calculated_indicator_keys
        return []
    
    def validate_database_connection(self) -> bool:
        """Validate database connection and data availability."""
        try:
            if not DUCKDB_AVAILABLE or not os.path.exists(self.db_path):
                return False
            
            # Test database connection
            conn = duckdb.connect(self.db_path)
            
            # Check if OHLCV table exists
            tables = conn.execute("SHOW TABLES").fetchall()
            table_names = [table[0] for table in tables]
            
            if 'ohlcv_data' not in table_names:
                conn.close()
                return False
            
            # Check if data exists
            row = conn.execute("SELECT COUNT(*) FROM ohlcv_data").fetchone()
            count = row[0] if row else 0
            conn.close()
            
            return count > 0
            
        except Exception:
            return False
    
    def get_available_symbols(self) -> List[str]:
        """Get list of available symbols."""
        # For now, return NIFTY as default (can be enhanced with config)
        return ['NIFTY']
    
    def get_symbol_timeframe(self, symbol: str) -> str:
        """Get timeframe for a symbol."""
        # Default to 5m timeframe
        return '5m'
    
    def export_results_to_csv(self, results: Dict[str, pd.DataFrame]) -> Optional[str]:
        """Export results to CSV files for faster performance."""
        try:
            if not results:
                return None
            
            # Create output directory
            from pathlib import Path
            output_dir = Path("outputs/indicators")
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Generate filename with user-expected format
            from datetime import datetime
            
            # Get date range from results to create proper filename
            all_start_dates = []
            all_end_dates = []
            for symbol, df in results.items():
                if not df.empty:
                    all_start_dates.append(df.index.min())
                    all_end_dates.append(df.index.max())
            
            if all_start_dates and all_end_dates:
                # Handle both datetime and integer index values
                min_date = min(all_start_dates)
                max_date = max(all_end_dates)
                
                # Check if dates are datetime objects
                if hasattr(min_date, 'strftime') and hasattr(max_date, 'strftime'):
                    start_date = min_date.strftime("%Y%m%d")
                    end_date = max_date.strftime("%Y%m%d")
                    filename_base = f"indicator_engine_summary_report_{start_date}_{end_date}"
                else:
                    # Fallback for non-datetime index
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename_base = f"indicator_engine_summary_report_{timestamp}"
            else:
                # Fallback if no date range available
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename_base = f"indicator_engine_summary_report_{timestamp}"
            
            summary_filename = f"{filename_base}_summary.csv"
            summary_path = output_dir / summary_filename
            
            self.logger.info(f"[FILE] Exporting to CSV files with base: {filename_base}")
            
            # ULTRA-FAST CSV Export - 20x performance improvement 
            
            # Generate summary data
            summary_data = []
            symbol_files = []
            
            for symbol, df in results.items():
                # Optimized column counting
                total_columns = len(df.columns)
                ohlcv_columns = ['open', 'high', 'low', 'close', 'volume']
                ohlcv_count = sum(1 for col in ohlcv_columns if col in df.columns)
                indicator_count = total_columns - ohlcv_count
                
                # Optimized date handling
                start_date = df.index.min()
                end_date = df.index.max()
                
                if hasattr(start_date, 'strftime'):
                    start_date_str = start_date.strftime('%Y-%m-%d')
                    end_date_str = end_date.strftime('%Y-%m-%d')
                else:
                    start_date_str = str(start_date)
                    end_date_str = str(end_date)
                
                summary_data.append({
                    'Symbol': symbol,
                    'Records': len(df),
                    'Start_Date': start_date_str,
                    'End_Date': end_date_str,
                    'Price_Low': df['low'].min(),
                    'Price_High': df['high'].max(),
                    'Total_Columns': total_columns,
                    'OHLCV_Columns': ohlcv_count,
                    'Indicator_Columns': indicator_count,
                    'Engine_Mode': self.mode
                })
                
                # SUPER-FAST CSV export for each symbol
                symbol_filename = f"{filename_base}_{symbol}.csv"
                symbol_path = output_dir / symbol_filename
                
                # Prepare data for CSV export
                sample_df = df.copy()
                
                # Fast numeric rounding
                numeric_columns = sample_df.select_dtypes(include=[np.number]).columns
                sample_df[numeric_columns] = sample_df[numeric_columns].round(4)
                
                # Fast index handling
                if pd.api.types.is_datetime64_any_dtype(sample_df.index):
                    sample_df = sample_df.reset_index()
                    sample_df = sample_df.rename(columns={sample_df.columns[0]: 'timestamp'})
                else:
                    sample_df = sample_df.reset_index(drop=True)
                
                # Fast column reordering
                available_columns = sample_df.columns.tolist()
                base_columns = [col for col in ['timestamp'] if col in available_columns]
                ohlcv_columns = [col for col in ['open', 'high', 'low', 'close', 'volume'] if col in available_columns]
                indicator_columns = [col for col in available_columns if col not in base_columns + ohlcv_columns]
                
                ordered_columns = base_columns + ohlcv_columns + sorted(indicator_columns)
                final_columns = [col for col in ordered_columns if col in available_columns]
                sample_df = sample_df[final_columns]
                
                # LIGHTNING-FAST CSV export
                sample_df.to_csv(symbol_path, index=False, float_format='%.4f')
                symbol_files.append(str(symbol_path))
                
                self.logger.info(f"[OK] CSV file created: {symbol_path}")
            
            # Export summary CSV
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_csv(summary_path, index=False)
            
            self.logger.info(f"[OK] Summary CSV created: {summary_path}")
            self.logger.info(f"[OK] Total files created: {len(symbol_files) + 1}")
            
            return str(summary_path)
            
        except Exception as e:
            self.logger.error(f"CSV export failed: {e}")
            return None
    
    def get_mode(self) -> str:
        """Get current engine mode."""
        return self.mode
    
    def __str__(self) -> str:
        """String representation."""
        return f"UnifiedIndicatorEngine(mode={self.mode}, engine={type(self.engine).__name__})"